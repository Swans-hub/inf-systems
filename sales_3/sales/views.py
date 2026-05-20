from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from io import BytesIO
from .models import FieldDefinition, DynamicRecord
from .forms import FieldDefinitionForm, DynamicRecordForm


# ========== Управление структурой (как в СУБД) ==========

def field_list(request):
    fields = FieldDefinition.objects.all().order_by('order')
    return render(request, 'sales/field_list.html', {'fields': fields})


def field_add(request):
    if request.method == 'POST':
        form = FieldDefinitionForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('field_list')
    else:
        form = FieldDefinitionForm()
    return render(request, 'sales/field_form.html', {'form': form, 'title': 'Добавить поле'})


def field_edit(request, pk):
    field = get_object_or_404(FieldDefinition, pk=pk)
    if request.method == 'POST':
        form = FieldDefinitionForm(request.POST, instance=field)
        if form.is_valid():
            form.save()
            return redirect('field_list')
    else:
        form = FieldDefinitionForm(instance=field)
    return render(request, 'sales/field_form.html', {'form': form, 'title': 'Редактировать поле'})


def field_delete(request, pk):
    field = get_object_or_404(FieldDefinition, pk=pk)
    if request.method == 'POST':
        field.delete()
        return redirect('field_list')
    return render(request, 'sales/field_confirm_delete.html', {'field': field})


# ========== Работа с данными ==========

def record_list(request):
    records = DynamicRecord.objects.all().order_by('-created_at')
    fields = FieldDefinition.objects.filter(is_active=True).order_by('order')
    return render(request, 'sales/record_list.html', {
        'records': records,
        'fields': fields,
    })


def record_add(request):
    if request.method == 'POST':
        form = DynamicRecordForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('record_list')
    else:
        form = DynamicRecordForm()
    return render(request, 'sales/record_form.html', {'form': form, 'title': 'Добавить запись'})


def record_edit(request, pk):
    record = get_object_or_404(DynamicRecord, pk=pk)
    if request.method == 'POST':
        form = DynamicRecordForm(request.POST, instance=record)
        if form.is_valid():
            form.save()
            return redirect('record_list')
    else:
        form = DynamicRecordForm(instance=record, initial=record.data)
    return render(request, 'sales/record_form.html', {'form': form, 'title': 'Редактировать запись'})


def record_delete(request, pk):
    record = get_object_or_404(DynamicRecord, pk=pk)
    if request.method == 'POST':
        record.delete()
        return redirect('record_list')
    return render(request, 'sales/record_confirm_delete.html', {'record': record})


# ========== Расчёт прибыли и экспорт ==========

def profit_report(request):
    records = DynamicRecord.objects.all()
    group_profit = {}

    for record in records:
        data = record.data
        group = data.get('product_group', 'Без группы')
        profit = record.get_profit()

        if group in group_profit:
            group_profit[group] += profit
        else:
            group_profit[group] = profit

    report_data = [{'group': k, 'profit': round(v, 2)} for k, v in group_profit.items()]
    total = sum(group_profit.values())

    return render(request, 'sales/profit_report.html', {
        'report_data': report_data,
        'total_profit': round(total, 2)
    })


def export_excel(request):
    records = DynamicRecord.objects.all()
    group_profit = {}

    for record in records:
        group = record.data.get('product_group', 'Без группы')
        profit = record.get_profit()
        if group in group_profit:
            group_profit[group] += profit
        else:
            group_profit[group] = profit

    wb = Workbook()
    ws = wb.active
    ws.title = "Прибыль по группам"

    ws['A1'] = "Группа товаров"
    ws['B1'] = "Прибыль, руб."

    row = 2
    for group, profit in group_profit.items():
        ws[f'A{row}'] = group
        ws[f'B{row}'] = profit
        row += 1

    ws[f'A{row}'] = "ИТОГО:"
    ws[f'B{row}'] = f"=SUM(B2:B{row - 1})"

    chart = BarChart()
    chart.title = "Прибыль по группам товаров"
    chart.x_axis.title = "Группа товаров"
    chart.y_axis.title = "Прибыль, руб."

    data = Reference(ws, min_col=2, min_row=1, max_row=row, max_col=2)
    categories = Reference(ws, min_col=1, min_row=2, max_row=row - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    ws.add_chart(chart, "D2")

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="profit_report.xlsx"'
    return response


def index(request):
    return render(request, 'sales/index.html')